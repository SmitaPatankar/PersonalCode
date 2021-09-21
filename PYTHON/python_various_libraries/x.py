from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# wb = load_workbook("C://SMITA PERSONAL REPOSITORY//GITHUB CODE//PYTHON//PYTHON_VARIOUS_LIBRARIES//dummy.xlsx")
# ws = wb.active
# print(ws["A1"].value)
# ws["A1"].value = "p"
# wb.save("temp.xlsx")
#
# print(wb.sheetnames)

# ws = wb["Sheet 1"]
#
# wb.create_sheet("test")
# print(wb.sheetnames)

# wb = Workbook()
# ws = wb.active
# ws.title = "Data"
# ws.append(["a", "b", "c", "d"])
# ws.append(["a", "b", "c", "d"])
# ws.append(["a", "b", "c", "d"])
# ws.append(["a", "b", "c", "d"])
# ws.append(["a", "b", "c", "d"])
# wb.save("temp.xlsx")

wb = load_workbook("C://SMITA PERSONAL REPOSITORY//GITHUB CODE//PYTHON//PYTHON_VARIOUS_LIBRARIES//dummy.xlsx")
ws = wb.active
for row in range(1, 11):
    for col in range(1, 5):
        char = get_column_letter(col)
        print(ws[char + str(row)].value)
wb.save("temp.xlsx")
ws.merge_cells("A1:D1")
ws.unmerge_cells("A1:D1")
ws.insert_rows(7)
ws.insert_cols(2)
ws.delete_cols(3)
ws.move_range("C1:d11", rows=2, columns=3)

from openpyxl.styles import Font

