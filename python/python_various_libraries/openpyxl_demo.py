from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# load workbook
wb = load_workbook("/PYTHON//PYTHON_VARIOUS_LIBRARIES//dummy.xlsx")
# get sheet names
print(wb.sheetnames)
# get active sheet
print(wb.active)
# get certain sheet
ws = wb[wb.sheetnames[0]]
print(ws)
# get cell value
print(ws["A1"].value)
# set cell value
'''ws["A1"].value = "names"'''
# create sheet
new_ws = wb.create_sheet("three")
# loop over rows
for row in range(1, 5):
    for col in range(1, 3):
        chr = get_column_letter(col)
        cell_identifier = chr + str(row)
        print(cell_identifier+"--->")
        print(ws[cell_identifier].value)
# merge cells
'''ws.merge("A1:D2")'''
# unmerge cells
'''ws.unmerge("A1:D2")'''
# insert blank row at
'''ws.insert_rows(1)'''
'''ws.insert_rows(1)'''
# delete row
'''ws.delete_rows(1)'''
# insert blank column at
'''ws.insert_cols(2)'''
# delete blank column at
'''ws.delete_cols(2)'''
# move data
'''ws.move_range("C1:D11", rows=2, cols=2)'''
# save workbook
wb.save("C://SMITA PERSONAL REPOSITORY//GITHUB CODE//PYTHON//PYTHON_VARIOUS_LIBRARIES//dummy.xlsx")

# ###################################################################################################

# create workbook
wb = Workbook()
# get active sheet
ws = wb.active
# set title of sheet
ws.title = "Data"
ws.append(["name", "city"])
ws.append(["smita", "dombivli"])
# save workbook
wb.save("C://SMITA PERSONAL REPOSITORY//GITHUB CODE//.temp//dummy.xlsx")

# ########################################################################

d = {
    "smita": {
        "surname": "patankar",
        "age": 29,
        "city": "mumbai"
    },
    "neha": {
        "surname": "patankar",
        "age": 31,
        "city": "sydney"
    }
}

wb = Workbook()
ws = wb.active
ws.title = "information"
headings = ["name"] + list(d["smita"].keys())
ws.append(headings)
for k, v in d.items():
    info = [k] + [v[heading] for heading in headings[1:]]
    ws.append(info)
ws["A1"].font = Font(bold=True, color="00FF8080")
wb.save("C://SMITA PERSONAL REPOSITORY//GITHUB CODE//.temp//dummy.xlsx")
